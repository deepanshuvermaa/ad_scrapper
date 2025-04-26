from flask import Flask, jsonify, request, send_file, render_template
from ads_scraper.scraper import extract_facebook_ad_data
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import logging
import traceback

# Set up logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger('flask_app')

app = Flask(__name__)

# Ensure downloads directory exists
os.makedirs('downloads', exist_ok=True)

def save_to_excel(ad_data, filename=None):  
    """Save ad data to Excel with embedded image URLs and video URLs."""
    if not ad_data:
        logger.warning("No ad data provided to save_to_excel")
        return None

    if not filename:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"facebook_ads_data_{timestamp}.xlsx"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Facebook Ads"

        headers = ['Library ID', 'Description', 'Image URL', 'Video URL', 'Backlink URL']
        ws.append(headers)

        column_widths = [33, 50, 65, 45, 45]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        for ad in ad_data:
            image_url = ad.get('Image URL', '')
            video_url = ad.get('Video URL', '')
            backlink_url = ad.get('Backlink URL', '')

            row = [
                ad.get('Library ID', ''),
                ad.get('Description', ''),
                image_url,
                video_url,
                backlink_url
            ]
            ws.append(row)

        # Save to downloads directory to ensure file is in a writable location
        file_path = os.path.join('downloads', filename)
        logger.info(f"Saving Excel file to {file_path}")
        wb.save(file_path)
        return file_path
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        logger.error(traceback.format_exc())
        return None


@app.route('/')
def home():
    logger.info("Home page requested")
    return render_template('index.html')


@app.route('/scrape', methods=['POST'])
def scrape_ads():
    try:
        # Get the URL parameter from the form
        url = request.form.get('url')
        logger.info(f"Scrape request received for URL: {url}")
        
        if not url:
            logger.warning("No URL provided in request")
            return jsonify({"error": "Please provide a URL parameter"}), 400

        # Extract ad data using the scraper
        logger.info(f"Starting extraction for URL: {url}")
        ad_data = extract_facebook_ad_data(url)
        
        if ad_data:
            logger.info(f"Extraction successful, found {len(ad_data)} ads")
            filename = save_to_excel(ad_data)
            
            if filename:
                logger.info(f"Returning Excel file: {filename}")
                return send_file(filename, as_attachment=True)
            else:
                logger.error("Failed to generate Excel file")
                return jsonify({"error": "Failed to generate Excel file"}), 500
        else:
            logger.warning("No data was extracted from the URL")
            return jsonify({"error": "No data was extracted"}), 500
    except Exception as e:
        # Log the full traceback for detailed error information
        logger.error(f"Unhandled exception in scrape_ads: {e}")
        logger.error(traceback.format_exc())
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    logger.info(f"Starting Flask app on port {port}")
    app.run(debug=True, host='0.0.0.0', port=port)